import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.filters import AutoFilter

def format_excel_file(file_path):
    """
    Format Excel file with bold headers, left alignment, auto-fit columns, and filters
    """
    try:
        workbook = load_workbook(file_path)
        
        sheet_names = ["SummaryTable", "Daily Summary", "Trip Summary", "Fleet Efficiency"]
    
        for sheet_name in sheet_names:
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Bold the first row, remove borders, and align left
                for cell in sheet[1]:
                    cell.font = Font(bold=True)
                    cell.border = None  # Remove any borders
                    cell.alignment = Alignment(horizontal='left')  # Left align
                
                # Auto-fit columns
                for column in sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    sheet.column_dimensions[column_letter].width = adjusted_width
                
                # Add auto-filter to the header row
                if sheet.max_row > 1 and sheet.max_column > 0:  # Only add filter if there's data
                    sheet.auto_filter.ref = f"A1:{sheet.cell(row=1, column=sheet.max_column).column_letter}1"
        
        workbook.save(file_path)
        workbook.close()
        
    except Exception as e:
        print(f"Error formatting Excel file: {e}")
        if 'workbook' in locals():
            workbook.close()